#!/usr/bin/env python3
"""
Create one unified Excel table from all scraper outputs.

Stacks rows from each source institution into a single table,
consolidates column names, cleans numeric values, and applies Excel formatting.
"""
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def parse_source_and_date(filename: str):
    """Extract source name and date from supported filename patterns."""
    date_match = DATE_RE.search(filename)
    file_date = date_match.group(1) if date_match else None

    if "_data_" in filename:
        source_name = filename.split("_data_")[0]
    else:
        # Legacy pattern: source_YYYY-MM-DD.xlsx
        source_name = re.sub(r"_\d{4}-\d{2}-\d{2}\.xlsx$", "", filename)

    return source_name, file_date


def institution_from_source(source_name: str) -> str:
    """Return the institution prefix from a source name (e.g. 'swedbank_pensions' -> 'swedbank')."""
    return source_name.split("_")[0]


def discover_latest_files_per_source():
    """
    Discover latest usable Excel file per source from both naming styles:
    - New: source_data_YYYY-MM-DD.xlsx
    - Legacy: source_YYYY-MM-DD.xlsx
    """
    candidates = []
    for path in Path(".").glob("*.xlsx"):
        lower = path.name.lower()
        if path.name.startswith("~$"):
            continue
        if "combined" in lower:
            continue
        if not DATE_RE.search(path.name):
            continue
        candidates.append(path)

    by_source = {}
    for path in candidates:
        source_name, _ = parse_source_and_date(path.name)
        if not source_name:
            continue

        current = by_source.get(source_name)
        if current is None or path.stat().st_mtime > current.stat().st_mtime:
            by_source[source_name] = path

    return by_source


def main():
    print("Discovering data files...")
    data_files = discover_latest_files_per_source()

    if not data_files:
        print("Error: No data files found. Run scrapers first.")
        sys.exit(1)

    print(f"Found {len(data_files)} data source(s):")
    for source, filepath in sorted(data_files.items()):
        print(f"  - {source}: {filepath.name}")

    # Read all source files, grouped by institution.
    by_institution = {}
    for source, filepath in sorted(data_files.items()):
        print(f"\nReading {source}...")
        df = pd.read_excel(filepath)
        print(f"  Loaded {len(df)} records, {len(df.columns)} columns")

        institution = institution_from_source(source)
        _, file_date = parse_source_and_date(filepath.name)

        if institution not in by_institution:
            by_institution[institution] = {"file_date": file_date, "dfs": []}
        by_institution[institution]["dfs"].append(df)

    # Within each institution, merge all its files on Fund name so each fund is one row.
    # Across institutions, stack rows.
    institution_frames = []
    for institution, info in sorted(by_institution.items()):
        dfs = info["dfs"]
        file_date = info["file_date"]

        if len(dfs) == 1:
            merged = dfs[0].copy()
        else:
            merged = dfs[0]
            for other in dfs[1:]:
                merged = merged.merge(other, on="Fund name", how="outer")

        institution_frames.append(merged)
        print(f"  Institution '{institution}': {len(merged)} funds")

    print("\nCombining all institutions into one table...")
    df_combined = pd.concat(institution_frames, ignore_index=True, sort=False)

    # Consolidate equivalent columns from different sources:
    # Date (Swedbank) -> Data
    if "Date" in df_combined.columns:
        if "Data" not in df_combined.columns:
            df_combined["Data"] = df_combined["Date"]
        else:
            df_combined["Data"] = df_combined["Data"].combine_first(df_combined["Date"])
        df_combined.drop(columns=["Date"], inplace=True)

    # GAV (Swedbank) -> Vieneto vertė
    if "GAV" in df_combined.columns:
        if "Vieneto vertė" not in df_combined.columns:
            df_combined["Vieneto vertė"] = df_combined["GAV"]
        else:
            df_combined["Vieneto vertė"] = df_combined["Vieneto vertė"].combine_first(df_combined["GAV"])
        df_combined.drop(columns=["GAV"], inplace=True)

    # Fondo dydis value (Swedbank) -> Grynieji aktyvai
    if "Fondo dydis value" in df_combined.columns:
        if "Grynieji aktyvai" not in df_combined.columns:
            df_combined["Grynieji aktyvai"] = df_combined["Fondo dydis value"]
        else:
            df_combined["Grynieji aktyvai"] = df_combined["Grynieji aktyvai"].combine_first(df_combined["Fondo dydis value"])
        df_combined.drop(columns=["Fondo dydis value"], inplace=True)

    # Sort by Fund name for readability
    if "Fund name" in df_combined.columns:
        df_combined.sort_values("Fund name", ignore_index=True, inplace=True)

    # Normalise Data column to YYYY-MM-DD (replace spaces/slashes with dashes)
    if "Data" in df_combined.columns:
        df_combined["Data"] = (
            df_combined["Data"]
            .astype(str)
            .str.strip()
            .str.replace(r"[\s/.]", "-", regex=True)
        )

    def clean_numeric(series):
        return pd.to_numeric(
            series.astype(str)
            .str.replace("EUR", "", regex=False)
            .str.replace(r"\s", "", regex=True)   # remove all whitespace (thousands sep)
            .str.replace(",", ".", regex=False)    # normalise decimal comma → dot
            .str.strip(),
            errors="coerce"
        )

    # Clean Vieneto vertė: strip "EUR", convert to numeric
    if "Vieneto vertė" in df_combined.columns:
        df_combined["Vieneto vertė"] = clean_numeric(df_combined["Vieneto vertė"])

    # Clean Grynieji aktyvai: strip "EUR", remove space thousands sep, convert to numeric
    if "Grynieji aktyvai" in df_combined.columns:
        df_combined["Grynieji aktyvai"] = clean_numeric(df_combined["Grynieji aktyvai"])

    print(f"  Combined: {len(df_combined)} rows, {len(df_combined.columns)} columns")

    # Use data date from the combined data for filename
    if 'Data' in df_combined.columns:
        unique_dates = df_combined['Data'].dropna().unique()
        if len(unique_dates) == 1 and unique_dates[0]:
            data_date = unique_dates[0]
        else:
            data_date = datetime.today().strftime("%Y-%m-%d")
    else:
        data_date = datetime.today().strftime("%Y-%m-%d")

    output_file = f"pension_data_combined_{data_date}.xlsx"

    # Rename column before writing
    df_combined.rename(columns={"Fund name": "Fondo pavadinimas"}, inplace=True)

    print(f"\nWriting to {output_file}...")
    df_combined.to_excel(output_file, index=False)

    # Apply formatting
    wb = load_workbook(output_file)
    ws = wb.active

    # Column widths
    ws.column_dimensions["A"].width = 40.12
    for col_letter in ["B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 21.5

    # Header row: bold, size 14, centered
    header_font = Font(bold=True, size=14)
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align

    wb.save(output_file)

    print(f"\n✅ Merged file created: {output_file}")
    print(f"   Rows: {len(df_combined)}")
    print(f"   Columns: {list(df_combined.columns)}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Error: {exc}")
        sys.exit(1)
